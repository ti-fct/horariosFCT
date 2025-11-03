import pandas as pd
import io
import os
import openpyxl
import logging
import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# --- Configurações ---
BASE_DIR = os.path.dirname(os.path.realpath(__file__))
CREDENTIALS_FILE = os.path.join(BASE_DIR, 'service_account.json')
LOG_FILE = os.path.join(BASE_DIR, 'logs.txt')
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
FILE_ID = '1uc8XtB7lc-rBNVuNoTaBfpR6gsLsPIuk'
CURRENT_EXCEL_PATH = os.path.join(BASE_DIR, 'planilha_horarios.xlsx')

# Obtenha o webhook do Google Chat das variáveis de ambiente
GOOGLE_CHAT_WEBHOOK_URL = os.environ.get('GOOGLE_CHAT_WEBHOOK')

# Configuração do logging
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

SHEET_NAMES = [
    "MINI-AUDITORIOS", "SALAS", "LABS INF-DES FCT", "LABS MAT-GEO",
    "LABS PROD", "LABS TRANSP"
]

def send_google_chat_notification(message):
    """Envia uma notificação para o Google Chat."""
    if not GOOGLE_CHAT_WEBHOOK_URL:
        print("Webhook do Google Chat não configurado. Pulando notificação.")
        logging.warning("Webhook do Google Chat não configurado.")
        return
    try:
        payload = {'text': message}
        requests.post(GOOGLE_CHAT_WEBHOOK_URL, json=payload)
    except Exception as e:
        print(f"Erro ao enviar mensagem para o Google Chat: {e}")
        logging.error(f"Erro ao enviar mensagem para o Google Chat: {e}")

def authenticate_google_drive():
    """Autentica com a API do Google Drive."""
    try:
        creds = service_account.Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=creds, cache_discovery=False)
        print("Autenticação com o Google Drive bem-sucedida.")
        logging.info("Autenticação com o Google Drive bem-sucedida.")
        return service
    except Exception as e:
        error_msg = f"Erro na autenticação com o Google Drive: {e}"
        print(error_msg)
        logging.error(error_msg)
        send_google_chat_notification(f"🚨 Erro na automação de horários: Falha na autenticação com o Google Drive. Verifique as credenciais. Detalhes: {e}")
        return None

def download_excel_from_drive(service, file_id, destination_path):
    """Baixa um arquivo do Google Drive."""
    try:
        request = service.files().get_media(fileId=file_id)
        with io.FileIO(destination_path, 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(f"Download: {int(status.progress() * 100)}%.")
        print(f"Arquivo baixado para: {destination_path}")
        logging.info(f"Arquivo baixado para: {destination_path}")
        return True
    except Exception as e:
        error_msg = f"Erro ao baixar o arquivo do Google Drive: {e}"
        print(error_msg)
        logging.error(error_msg)
        send_google_chat_notification(f"🚨 Erro na automação de horários: Falha ao baixar a planilha. Detalhes: {e}")
        return False

def process_excel_sheet(sheet):
    """Processa uma única aba da planilha para extrair os dados das salas."""
    all_room_data = []
    sheet_data = [[str(cell if cell is not None else '').strip() for cell in row] for row in sheet.iter_rows(values_only=True)]

    room_indices = [0, 16, 32, 48, 64, 80, 96, 112]
    current_room_name = None
    start_data_row = -1

    for i, row_values in enumerate(sheet_data):
        if i in room_indices:
            line = ','.join(row_values)
            if " - SALA" in line:
                current_room_name = line.split(" - SALA")[1].strip().split(',')[0].strip()
                start_data_row = i + 2

        if current_room_name and i >= start_data_row and row_values[0]:
            row_dict = {
                "Horário": row_values[0],
                "Segunda": row_values[1] if len(row_values) > 1 else "",
                "Terça": row_values[2] if len(row_values) > 2 else "",
                "Quarta": row_values[3] if len(row_values) > 3 else "",
                "Quinta": row_values[4] if len(row_values) > 4 else "",
                "Sexta": row_values[5] if len(row_values) > 5 else "",
                "Sábado": row_values[6] if len(row_values) > 6 else "",
                "Sala": current_room_name
            }
            all_room_data.append(row_dict)

    return pd.DataFrame(all_room_data)

def format_csv_to_js(csv_path, js_output_path):
    """Converte o arquivo CSV para um array JavaScript."""
    try:
        df = pd.read_csv(csv_path).fillna('')
        data_list = df.values.tolist()

        with open(js_output_path, 'w', encoding='utf-8') as js_file:
            js_file.write('var dados = [\n')
            for i, row in enumerate(data_list):
                # Processa cada item da linha para escapar aspas simples e barras invertidas
                processed_row = []
                for item in row:
                    # 1. Converte para string
                    # 2. Substitui barras invertidas por duas barras invertidas (escapa a própria barra)
                    # 3. Substitui aspas simples por uma barra invertida e uma aspa simples (escapa a aspa)
                    escaped_item = str(item).replace('\\', '\\\\').replace("'", "\\'")
                    processed_row.append(f"'{escaped_item}'")
                
                # Junta os itens processados
                formatted_row = ', '.join(processed_row)
                
                js_file.write(f"  [{formatted_row}]")
                if i < len(data_list) - 1:
                    js_file.write(',\n')
                else:
                    js_file.write('\n')
            js_file.write('];\n')
        print("Arquivo dados.js gerado com sucesso.")
        logging.info("Arquivo dados.js gerado com sucesso.")
    except Exception as e:
        error_msg = f"Erro ao gerar o arquivo JS: {e}"
        print(error_msg)
        logging.error(error_msg)
        send_google_chat_notification(f"🚨 Erro na automação de horários: Falha ao converter CSV para JS. Detalhes: {e}")
        
def main():
    """Função principal para orquestrar o processo."""
    drive_service = authenticate_google_drive()
    if not drive_service:
        return

    if not download_excel_from_drive(drive_service, FILE_ID, CURRENT_EXCEL_PATH):
        return

    try:
        workbook = openpyxl.load_workbook(CURRENT_EXCEL_PATH)
        combined_df = pd.DataFrame()

        for sheet_name in SHEET_NAMES:
            if sheet_name in workbook.sheetnames:
                print(f"Processando aba: {sheet_name}")
                sheet = workbook[sheet_name]
                df_sheet = process_excel_sheet(sheet)
                combined_df = pd.concat([combined_df, df_sheet], ignore_index=True)
            else:
                logging.warning(f"Aba '{sheet_name}' não encontrada. Ignorando.")

        desired_columns = ["Horário", "Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Sala"]
        combined_df = combined_df[desired_columns]

        # Caminhos relativos à raiz do repositório
        csv_output_path = os.path.join(BASE_DIR, 'Horarios_Salas_Combinados.csv')
        js_output_path = os.path.join(BASE_DIR, '..', 'js', 'dados.js')

        combined_df.to_csv(csv_output_path, index=False)
        print(f"Dados combinados salvos em {csv_output_path}")
        logging.info(f"Dados combinados salvos em {csv_output_path}")

        format_csv_to_js(csv_output_path, js_output_path)

        send_google_chat_notification("✅ Horários da FCT atualizados com sucesso!")

    except Exception as e:
        error_msg = f"Erro ao processar o arquivo Excel: {e}"
        print(error_msg)
        logging.error(error_msg, exc_info=True)
        send_google_chat_notification(f"🚨 Erro na automação de horários: Falha ao processar a planilha. Detalhes: {e}")
    finally:
        # Limpa o arquivo baixado
        if os.path.exists(CURRENT_EXCEL_PATH):
            os.remove(CURRENT_EXCEL_PATH)
            print(f"Arquivo temporário '{CURRENT_EXCEL_PATH}' removido.")


if __name__ == "__main__":
    main()