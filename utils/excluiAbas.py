import openpyxl

def excluir_abas(arquivo, abas_a_excluir):
  
  # Carrega o arquivo Excel
  wb = openpyxl.load_workbook(arquivo)

  # Cria uma lista com os nomes das abas a serem mantidas
  abas_a_manter = [wb.sheetnames[i] for i in range(len(wb.sheetnames)) if i not in abas_a_excluir]

  # Remove as abas indesejadas
  for aba in wb.sheetnames:
    if aba not in abas_a_manter:
      wb.remove(wb[aba])

  # Salva o arquivo com as alterações
  wb.save(arquivo)

# Exemplo de uso:
arquivo_excel = '2024-2-Disciplinas e Salas - FCT.xlsx'  # Substitua pelo caminho do seu arquivo
abas_para_excluir = [6, 7, 8]  # Ajustar os índices de acordo com as abas desejadas

excluir_abas(arquivo_excel, abas_para_excluir)