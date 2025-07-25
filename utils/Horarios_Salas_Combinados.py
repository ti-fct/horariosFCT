import pandas as pd
import io
import os
import hashlib
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl
import Send_to_Github

# --- Configurações ---
SERVICE_ACCOUNT_FILE = 'service_account.json'
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
FILE_ID = '1kcimnmHy4w7gY1nZu37lx6mlditiu4UC'
SHEET_NAMES = [
    "MINI-AUDITORIOS",
    "SALAS",
    "LABS INF-DES FCT",
    "LABS MAT-GEO",
    "LABS PROD",
    "LABS TRANSP"
]
# Caminhos para o arquivo atual e o arquivo antigo
CURRENT_EXCEL_PATH = 'planilha_horarios.xlsx'
OLD_EXCEL_PATH = 'planilha_horarios_antiga.xlsx'

# --- Funções de Autenticação

def authenticate_google_drive():
    try:
        credentials = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=credentials)
        print("Autenticação com o Google Drive bem-sucedida.")
        return service
    except Exception as e:
        print(f"Erro na autenticação com o Google Drive: {e}")
        return None
    
# --- Função de Download do Arquivo Excel ---

def download_excel_from_drive(service, file_id, destination_path):
    try:
        request = service.files().get_media(fileId=file_id)
        fh = io.FileIO(destination_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print(f"Download: {int(status.progress() * 100)}%.")
        print(f"Arquivo baixado para: {destination_path}")
        return True
    except Exception as e:
        print(f"Erro ao baixar o arquivo do Google Drive: {e}")
        return False

# --- Função de Comparação de Arquivos ---

def files_are_different(file_path1, file_path2):
    """
    Compara dois arquivos XLSX calculando seus hashes SHA-256.
    Retorna True se os arquivos forem diferentes, False se forem iguais.
    """
    if not os.path.exists(file_path1) or not os.path.exists(file_path2):
        return True  # Se um dos arquivos não existe, considera como diferente

    def calculate_file_hash(file_path):
        sha256_hash = hashlib.sha256()
        with open(file_path, 'rb') as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()

    hash1 = calculate_file_hash(file_path1)
    hash2 = calculate_file_hash(file_path2)
    return hash1 != hash2

# --- Função de Processamento de Dados ---

def process_excel_sheet(sheet, sheet_name):
    all_room_data = []
    sheet_data = []
    for row in sheet.iter_rows(values_only=True):
        sheet_data.append([str(cell if cell is not None else '').strip() for cell in row])

    room_indices = [0, 16, 32, 48, 64, 80, 96, 112]
    current_room_name = None
    start_data_row = -1

    for i, row_values in enumerate(sheet_data):
        line = ','.join(row_values)
        for idx in room_indices:
            if i == idx:
                if " - SALA" in line:
                    current_room_name = line.split(" - SALA")[1].strip().split(',')[0].strip()
                    start_data_row = i + 2
                break
        else:
            pass

        if current_room_name and i >= start_data_row:
            if len(row_values) > 1 and any(cell for cell in row_values[1:]):
                time = row_values[0]
                if not time:
                    continue
                row_dict = {
                    "Horário": time,
                    "Segunda": row_values[1] if len(row_values) > 1 else "",
                    "Terça": row_values[2] if len(row_values) > 2 else "",
                    "Quarta": row_values[3] if len(row_values) > 3 else "",
                    "Quinta": row_values[4] if len(row_values) > 4 else "",
                    "Sexta": row_values[5] if len(row_values) > 5 else "",
                    "Sábado": row_values[6] if len(row_values) > 6 else "",
                    "Sala": current_room_name
                }
                all_room_data.append(row_dict)

    return pd.DataFrame(all_room_data)

# --- Lógica Principal ---

if __name__ == "__main__":
    # 1. Autenticar com o Google Drive
    drive_service = authenticate_google_drive()

    if drive_service:
        # 2. Baixar o novo arquivo Excel do Google Drive
        if download_excel_from_drive(drive_service, FILE_ID, CURRENT_EXCEL_PATH):
            # 3. Comparar com o arquivo antigo
            if files_are_different(CURRENT_EXCEL_PATH, OLD_EXCEL_PATH):
                print("Alterações detectadas. Processando o novo arquivo...")
                combined_df = pd.DataFrame()

                try:
                    # 4. Carregar o arquivo Excel com openpyxl
                    workbook = openpyxl.load_workbook(CURRENT_EXCEL_PATH)
                    print(f"Arquivo Excel '{CURRENT_EXCEL_PATH}' carregado com sucesso.")

                    # 5. Processar cada aba especificada
                    for sheet_name in SHEET_NAMES:
                        if sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            print(f"Processando aba: {sheet_name}")
                            df_sheet = process_excel_sheet(sheet, sheet_name)
                            combined_df = pd.concat([combined_df, df_sheet], ignore_index=True)
                        else:
                            print(f"Aba '{sheet_name}' não encontrada no arquivo Excel. Ignorando.")

                    # 6. Definir a ordem desejada das colunas
                    desired_columns = ["Horário", "Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Sala"]
                    existing_desired_columns = [col for col in desired_columns if col in combined_df.columns]
                    combined_df = combined_df[existing_desired_columns]

                    # 7. Salvar o DataFrame combinado em um arquivo CSV
                    output_csv_file = "Horarios_Salas_Combinados.csv"
                    combined_df.to_csv(output_csv_file, index=False)

                    print(f"\nDados combinados salvos em {output_csv_file}")
                    
                    # Envia para o GitHub
                    Send_to_Github.upload_to_github(output_csv_file, 'ti-fct', 'horariosFCT', 'main', 'utils/'+output_csv_file)
                    
                    # 8. Substituir o arquivo antigo pelo novo
                    if os.path.exists(OLD_EXCEL_PATH):
                        os.remove(OLD_EXCEL_PATH)
                    os.rename(CURRENT_EXCEL_PATH, OLD_EXCEL_PATH)
                    print(f"Arquivo antigo atualizado: {OLD_EXCEL_PATH}")

                except Exception as e:
                    print(f"Erro ao processar o arquivo Excel: {e}")
                    # Não remove o arquivo baixado em caso de erro para evitar perda de dados
            else:
                print("Nenhuma alteração detectada. O script não será executado.")
                # Remove o arquivo baixado, já que não houve alterações
                if os.path.exists(CURRENT_EXCEL_PATH):
                    os.remove(CURRENT_EXCEL_PATH)
                    print(f"Arquivo temporário '{CURRENT_EXCEL_PATH}' removido.")
        else:
            print("Não foi possível baixar o arquivo Excel. O processamento foi interrompido.")
    else:
        print("Não foi possível autenticar com o Google Drive. O processamento foi interrompido.")